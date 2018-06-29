from apps.sales.models import Sale
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import json

REPORTE_VENTAS = {
    0: 'Ventas Contado',
    1: 'Ventas Crédito',
    2: 'Abonos',
    3: 'Notas Crédito',
    4: 'Productos',
    5: 'Adelantos',
    6: 'Cierres de Taller',
}

COMPANY_NAME = 'Repuestos Juanca'

def createGenSalesReport(start, end, day, month, year):
    print("Create gen sales report entry point")

    #variable to track the time covered
    report_coverage = ''

    #declare the names of the worksheets to be created
    wb = Workbook()
    
    today = date.today()


    report_period_type = None
    if start!='' or end!='':
        report_period_type='date_range'
    elif day!='':
        report_period_type ='day'
    elif month!='':
        report_period_type = "month"
    elif year != '':
        report_period_type = "year"
    
    target_sales = None
    if report_period_type == 'day':
        day_target =  None
        month_target = None
        year_target = None

        if day == 'THIS':
            target_sales =  Sale.objects.filter(created__lt=today+timedelta(days=1), 
                created__gt=today-timedelta(days=1))
            report_coverage = today
        else:
            day_target = int(day)
            if month == '':
                month_target = today.month
            else:
                month_target = int(month)
            if year == '':
                year_target = today.year
            else:
                year_target = int(year)
            target_date = datetime(year_target, month_target, day_target)
            report_coverage = target_date
            target_sales = Sale.objects.filter(created__lt = target_date+timedelta(days=1),
                created__gt=target_date-timedelta(days=1))
    elif report_period_type == "month":
        target_year = None

        if year == '':
            target_year = today.year
        else:
            target_year = int(year)
        target_month = None
        if month == 'THIS':
            target_month = today.month
        else:
            target_month = int(month)
        report_coverage = "Mes: {}, Año: {}".format(target_month, target_year)
        target_sales = Sale.objects.filter(created__month = target_month, created__year =target_year)
    elif report_period_type == "year":
        target_year = None
        if year == 'THIS':
            target_year = today.year
        else:
            target_year = int(year)
        report_coverage = "Año: {}".format(target_year)
        target_sales = Sale.objects.filter(created__year=target_year)
    elif report_period_type == "date_range":
        #expected date  is YYYY-MONTH-DAY
        start_year, start_month, start_day = start.split('-')
        start_year = int(start_year)
        start_month = int(start_month)
        start_day = int(start_day)

        end_year, end_month, end_day = end.split('-')
        end_year = int(end_year)
        end_month = int(end_month)
        end_day = int(end_day)

        start_date = datetime(start_year, start_month, start_day)
        end_date = datetime(end_year, end_month, end_day)

        report_coverage = "Inicio: {}, Fin: {}".format(start_date, end_date)
        target_sales =  Sale.objects.filter(created__lt=end_date, 
            created__gt=start_date)

    #separate the sales into the different reports that belong to
    cash_sales = []
    cred_sales = []

    for sale in target_sales:
        if sale.sale_type == 'CASH':
            cart_data =  json.loads(sale.cart)
            cash_sales.append((sale, cart_data))
        else:
            cart_data =  json.loads(sale.cart)
            cred_sales.append((sale, cart_data))
    

    createCashSalesReport(cash_sales, wb, report_coverage)
    createCreditSalesReport(cred_sales, wb, report_coverage)
    createProductMovementReport(cash_sales, cred_sales, wb, report_coverage)
    return wb

def createProductMovementReport(target_sales_cash, target_sales_cred, wb, period, sheet_index = 4):

    current_row = 1
    current_col = 1
    header_labels = ['Código', 'Descripción', 'Unidades', 'Subtotal', 'Descuento', 'Impuesto', 'Total']

    ws = wb.create_sheet(title=REPORTE_VENTAS[sheet_index])

    #add a simple title
    ws.cell(row=current_row, column=current_col, value=COMPANY_NAME)
    current_row+=1
    #write the date
    now = datetime.now()
    report_generation = "Fecha generación: {}".format(now)
    ws.cell(row=current_row, column=current_col, value=report_generation)
    current_row+=1
    #add the perdiod covered by the report
    report_generation = "Periódo cubierto: {}".format(period)
    ws.cell(row=current_row, column=current_col, value=report_generation)
    current_row+=1
    #write the report type
    ws.cell(row=current_row, column=current_col, value="REPORTE FACTURACIÓN DE CRÉDITO")
    current_row+=2

    #write the header
    for i in range(0,len(header_labels)):
        temp_cell = ws.cell(row=current_row, column=current_col+i, value=header_labels[i])
        temp_cell.font = Font(bold=True, name="Tahoma", size=8)
    #reset the col position
    current_col = 1
    current_row+=1

    data_body = []
    total_subtotal = 0
    total_discount = 0
    total_tax = 0
    total_total = 0
    for element in (target_sales_cash + target_sales_cred):
        for item in element[1]['cartItems']: #each cart can have many items
            temp_row = []
            #get item code
            temp_row.append(item['product']['code'])
            #get item description
            temp_row.append(item['product']['description'])
            #get item qty
            temp_row.append(round(item['qty'], 2))
            #get item subtotal
            temp_subtotal = item['subTotalNoDiscount']
            total_subtotal += temp_subtotal
            temp_row.append(round(temp_subtotal, 2))
            #get item discount
            temp_discount = item['subTotalNoDiscount']-item['subtotal']
            total_discount += temp_discount
            temp_row.append(round(temp_discount, 2))
            #get item tax
            temp_tax = item['totalWithIv']-item['subtotal']
            total_tax += temp_tax
            temp_row.append(round(total_tax, 2))
            #get item total
            temp_total = item['totalWithIv']
            total_total += temp_total
            temp_row.append(round(temp_total, 2))
            data_body.append(temp_row)

    for data_row in data_body:
        for i in range(0,len(data_row)):
            temp_cell = ws.cell(row=current_row, column=current_col+i, value=data_row[i])
        current_row += 1
        current_col = 1
        
    current_col = 3
    current_row += 2
    totals_legend_cell = ws.cell(row=current_row, column=current_col, value="Totales-->")
    totals_legend_cell.font = Font(bold=True, name="Tahoma", size=8)
    current_col+=1

    #add the subtotal
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_subtotal,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1

    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_discount,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1

    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_tax,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1
    
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_total,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1

    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

def createCashSalesReport(target_sales, wb, period, sheet_index = 0):

    current_row = 1
    current_col = 1
    header_labels = ['# Factura', 'Cliente', 'Subtotal', 'Descuento', 'Impuestos', 'Total', 'Fecha Venta' ]

    wb._active_sheet_index =  sheet_index
    #get the first sheet, created per default
    ws = wb.active
    #change the title of the sheet
    ws.title = REPORTE_VENTAS[0]

    #add a simple title
    ws.cell(row=current_row, column=current_col, value=COMPANY_NAME)
    current_row+=1
    #write the date
    now = datetime.now()
    report_generation = "Fecha generación: {}".format(now)
    ws.cell(row=current_row, column=current_col, value=report_generation)
    current_row+=1
    #add the perdiod covered by the report
    report_generation = "Periódo cubierto: {}".format(period)
    ws.cell(row=current_row, column=current_col, value=report_generation)
    current_row+=1
    #write the report type
    ws.cell(row=current_row, column=current_col, value="REPORTE FACTURACIÓN DE CONTADO")
    current_row+=2
    #write the header
    for i in range(0,len(header_labels)):
        temp_cell = ws.cell(row=current_row, column=current_col+i, value=header_labels[i])
        temp_cell.font = Font(bold=True, name="Tahoma", size=8)
    #reset the col position
    current_col = 1
    current_row+=1
    data_body = []
    total_subtotal = 0
    total_discount = 0
    total_tax = 0
    total_total = 0
    for element in target_sales:
        temp_row = []
        #add invoice number
        temp_row.append(element[0].consecutive)
        #get client data
        client = json.loads(element[0].client)
        temp_row.append("{} {}".format(client['name'], client['last_name']) )
        #get invoice subtotal
        tempt_subtotal = element[1]['cartSubtotalNoDiscount']
        total_subtotal += tempt_subtotal
        temp_row.append(round(tempt_subtotal,2))
        #get the discount
        temp_discount = element[1]['discountTotal']
        total_discount += temp_discount
        temp_row.append(round(temp_discount,2))
        #get the total taxes
        temp_taxes = element[1]['cartTaxes']
        total_tax += temp_taxes
        temp_row.append(round(temp_taxes,2))
        #get the total
        temp_total = element[0].sale_total
        total_total += temp_total
        temp_row.append(round(temp_total,2))
        #get the sale time
        temp_row.append(element[0].created)
        data_body.append(temp_row)

    for data_row in data_body:
        for i in range(0,len(data_row)):
            temp_cell = ws.cell(row=current_row, column=current_col+i, value=data_row[i])
        current_row += 1
        current_col = 1
    #calculate the totals
    current_col = 2
    current_row += 2
    totals_legend_cell = ws.cell(row=current_row, column=current_col, value="Totales-->")
    totals_legend_cell.font = Font(bold=True, name="Tahoma", size=8)
    current_col+=1
    #add the subtotal
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_subtotal,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1

    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_discount,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1

    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_tax,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1
    
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_total,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1


    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width


def createCreditSalesReport(target_sales, wb, period, sheet_index = 1):

    current_row = 1
    current_col = 1
    header_labels = ['# Factura', 'Cliente', 'Subtotal', 'Descuento', 'Impuestos', 'Total', 'Fecha Venta' ]    

    ws = wb.create_sheet(title=REPORTE_VENTAS[sheet_index])

    #add a simple title
    ws.cell(row=current_row, column=current_col, value=COMPANY_NAME)
    current_row+=1
    #write the date
    now = datetime.now()
    report_generation = "Fecha generación: {}".format(now)
    ws.cell(row=current_row, column=current_col, value=report_generation)
    current_row+=1
    #add the perdiod covered by the report
    report_generation = "Periódo cubierto: {}".format(period)
    ws.cell(row=current_row, column=current_col, value=report_generation)
    current_row+=1
    #write the report type
    ws.cell(row=current_row, column=current_col, value="REPORTE FACTURACIÓN DE CRÉDITO")
    current_row+=2

    #write the header
    for i in range(0,len(header_labels)):
        temp_cell = ws.cell(row=current_row, column=current_col+i, value=header_labels[i])
        temp_cell.font = Font(bold=True, name="Tahoma", size=8)
    #reset the col position
    current_col = 1
    current_row+=1
    data_body = []
    total_subtotal = 0
    total_discount = 0
    total_tax = 0
    total_total = 0

    for element in target_sales:
        temp_row = []
        #add invoice number
        temp_row.append(element[0].consecutive)
        #get client data
        client = json.loads(element[0].client)
        temp_row.append("{} {}".format(client['name'], client['last_name']) )
        #get invoice subtotal
        tempt_subtotal = element[1]['cartSubtotalNoDiscount']
        total_subtotal += tempt_subtotal
        temp_row.append(round(tempt_subtotal,2))
        #get the discount
        temp_discount = element[1]['discountTotal']
        total_discount += temp_discount
        temp_row.append(round(temp_discount,2))
        #get the total taxes
        temp_taxes = element[1]['cartTaxes']
        total_tax += temp_taxes
        temp_row.append(round(temp_taxes,2))
        #get the total
        temp_total = element[0].sale_total
        total_total += temp_total
        temp_row.append(round(temp_total,2))
        #get the sale time
        temp_row.append(element[0].created)
        data_body.append(temp_row)

    for data_row in data_body:
        for i in range(0,len(data_row)):
            temp_cell = ws.cell(row=current_row, column=current_col+i, value=data_row[i])
        current_row += 1
        current_col = 1
    #calculate the totals
    current_col = 2
    current_row += 2
    totals_legend_cell = ws.cell(row=current_row, column=current_col, value="Totales-->")
    totals_legend_cell.font = Font(bold=True, name="Tahoma", size=8)
    current_col+=1
    #add the subtotal
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_subtotal,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1

    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_discount,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1

    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_tax,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1
    
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_total,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1


    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width