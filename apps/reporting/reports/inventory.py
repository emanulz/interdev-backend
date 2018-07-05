from apps.products.models import Product
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import json

from .helpers import getCompanyName, convertToLocalTimezone, adjustSheetColsWidth

VALORACION_INVENTARIO = {
    0: 'Valoración Inventario',
}


def createInventoryValueReport(target_warehouses, include_inactive=False):
    '''
    Totalizes the value of all items in the inventory
    '''
    print('Inventory value report')
    print(target_warehouses)

    target_prods = Product.objects.filter(is_active=not include_inactive).filter(inventory_enabled=True)
    wb = Workbook()
    createInvValueSheet(wb, target_prods)
    return wb


def createInvValueSheet(wb, target_prods, sheet_index = 0):

    COMPANY_NAME = getCompanyName()

    current_row = 1
    current_col = 1
    header_labels = ['Código', 'Descripción', 'Unidades', 'Costo', 'Total']

    wb._active_sheet_index =  sheet_index

    ws = wb.active
    #change the title of the sheet
    ws.title = VALORACION_INVENTARIO[0]
    #add a simple title
    ws.cell(row=current_row, column=current_col, value=COMPANY_NAME)
    #write the date
    now = datetime.now()
    report_generation = "Fecha generación: {}".format(now)
    ws.cell(row=current_row, column=current_col, value=report_generation)
    current_row+=1
    #write the report type
    ws.cell(row=current_row, column=current_col, value="REPORTE VALORACIÓN INVENTARIO")
    current_row+=2
    #write the header
    for i in range(0,len(header_labels)):
        temp_cell = ws.cell(row=current_row, column=current_col+i, value=header_labels[i])
        temp_cell.font = Font(bold=True, name="Tahoma", size=8)
    #reset the col position
    current_col = 1
    current_row+=1
    data_body = []
    total = 0
    for element in target_prods:
        temp_row = []
        #get the element code
        temp_row.append(element.code)
        #get the element description
        temp_row.append(element.description)
        #load the inventory and calculate how many units are available
        temp_inv= json.loads(element.inventory_existent)
        temp_units = 0
        #for key in temp_inv.keys():
        temp_units += float(temp_inv['total'])
        temp_row.append(round(temp_units, 2))
        #get the cost        
        temp_cost = element.cost
        temp_row.append(round(temp_cost, 2))
        #get the value
        temp_value =  temp_cost *  temp_units
        temp_row.append(round(temp_value,2))
        #track the total inventory value
        total += temp_value
        data_body.append(temp_row)

    for data_row in data_body:
        for i in range(0,len(data_row)):
            temp_cell = ws.cell(row=current_row, column=current_col+i, value=data_row[i])
        current_row += 1
        current_col = 1

    current_col = 4
    current_row += 2
    totals_legend_cell = ws.cell(row=current_row, column=current_col, value="Total-->")
    totals_legend_cell.font = Font(bold=True, name="Tahoma", size=8)
    totals_legend_cell.number_format = '#,##0.00₡' 
    current_col+=1
    
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1

    adjustSheetColsWidth(ws)

