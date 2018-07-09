from purchases.models import Purchase
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import json
from datetime import datetime, timedelta, date
from apps.utils.serializers import UserSerialiazer
from .helpers import getCompanyName, convertToLocalTimezone, adjustSheetColsWidth, getCreditExpiryDate
from django.contrib.auth.models import User

REPORTE_COMPRAS = {
    0: 'Compras Crédito',
    1: 'Compras Contado',
}

def createGenPurchasesreport(start, end, day, month, year):
    print("Create purchases report entry point")

    #variable to track the time covered
    report_coverage = ''

    #declare the names of the worksheets to be created
    wb = Workbook()
    
    today = date.today()

    report_period_type = None

    if start!='' or end!='':
        report_period_type='date_range'
    elif day!='':
        report_period_type ='day'
    elif month!='':
        report_period_type = "month"
    elif year != '':
        report_period_type = "year"

    target_purchases = None

    if report_period_type == 'day':
        day_target =  None
        month_target = None
        year_target = None

        if day == 'THIS':
            target_purchases =  Purchase.objects.filter(created__lt=today+timedelta(days=1), 
                created__gt=today-timedelta(days=1))
            report_coverage = today
        else:
            day_target = int(day)
            if month == '':
                month_target = today.month
            else:
                month_target = int(month)
            if year == '':
                year_target = today.year
            else:
                year_target = int(year)
            target_date = datetime(year_target, month_target, day_target)
            report_coverage = target_date

            target_purchases = Purchase.objects.filter(created__lt = target_date+timedelta(days=1),
                created__gt=target_date-timedelta(days=1))

    elif report_period_type == "month":
        target_year = None

        if year == '':
            target_year = today.year
        else:
            target_year = int(year)
        target_month = None
        if month == 'THIS':
            target_month = today.month
        else:
            target_month = int(month)
        report_coverage = "Mes: {}, Año: {}".format(target_month, target_year)
        target_purchases = Purchase.objects.filter(created__month = target_month, created__year =target_year)
    elif report_period_type == "year":
        target_year = None
        if year == 'THIS':
            target_year = today.year
        else:
            target_year = int(year)
        report_coverage = "Año: {}".format(target_year)
        target_purchases = Purchase.objects.filter(created__year=target_year)
    elif report_period_type == "date_range":
        #expected date  is YYYY-MONTH-DAY
        start_year, start_month, start_day = start.split('-')
        start_year = int(start_year)
        start_month = int(start_month)
        start_day = int(start_day)

        end_year, end_month, end_day = end.split('-')
        end_year = int(end_year)
        end_month = int(end_month)
        end_day = int(end_day)

        start_date = datetime(start_year, start_month, start_day)
        end_date = datetime(end_year, end_month, end_day)

        report_coverage = "Inicio: {}, Fin: {}".format(start_date, end_date)
        target_purchases =  Purchase.objects.filter(created__lt=end_date, 
            created__gt=start_date)

    cash_purchases = []
    credit_purchases = []

    for purchase in target_purchases:
        try:
            cart_data =  json.loads(purchase.cart)
        except:
            print("Purchase broken --> ", purchase.consecutive)
            continue
        if purchase.purchase_type == 'CASH':
            cash_purchases.append((purchase, cart_data))
        else:
            credit_purchases.append((purchase, cart_data))

    createCreditPurchasesReport(credit_purchases, wb, report_coverage)
    createCashPurchasesReport(cash_purchases, wb, report_coverage)
    return wb

def createCreditPurchasesReport(purchases, wb, period, sheet_index=0):
    COMPANY_NAME = getCompanyName()
    current_row = 1
    current_col = 1
    header_labels = ['# Consecutivo', '# Factura', 'Fecha Factura', 'Proveedor', 'Subtotal', 'Descuento', 'Transporte', 'Impuesto', 'Total', 'Plazo', 'Vencimiento', 'Fecha Ingreso', 'Usuario' ] 
    ws = None
    if sheet_index == 0:
        wb._active_sheet_index = sheet_index
        #get the first sheet, created per default
        ws = wb.active
        #change the title of the sheet
        ws.title = REPORTE_COMPRAS[0]
    else:
        ws = wb.create_sheet(title=REPORTE_COMPRAS[sheet_index])

    #add a simple title
    ws.cell(row=current_row, column=current_col, value=COMPANY_NAME)
    current_row+=1
    #write the date
    now = datetime.now()
    report_generation = "Fecha generación: {}".format(now)
    ws.cell(row=current_row, column=current_col, value=report_generation)
    current_row+=1
    #add the perdiod covered by the report
    report_generation = "Periódo cubierto: {}".format(period)
    ws.cell(row=current_row, column=current_col, value=report_generation)
    current_row+=1
    #write the report type
    ws.cell(row=current_row, column=current_col, value="REPORTE COMPRAS A CRÉDITO")
    current_row+=2

    #write the header
    for i in range(0,len(header_labels)):
        temp_cell = ws.cell(row=current_row, column=current_col+i, value=header_labels[i])
        temp_cell.font = Font(bold=True, name="Tahoma", size=8)
    #reset the col position
    current_col = 1
    current_row+=1
    
    #variable to hold the data of the sheet body
    data_body = []
    
    #track totals for the different columns
    total_subtotal = 0
    total_discount = 0
    total_transport = 0
    total_tax = 0
    total_total = 0
    
    #each element is a tupple with the first index being the purchase
    #and the second on the parsed json cart
    for element in purchases:
        temp_row = []
        #get the internal system consecutive
        temp_row.append(element[0].consecutive)
        #get the invoice number as received
        temp_row.append(element[0].invoice_number)
        #get the date on which the invoice was emitted
        temp_row.append(element[0].invoice_date)
        #get the supplier
        supplier = json.loads(element[0].supplier)
        temp_row.append("{}-{}".format(supplier['code'], supplier['name']))
        #get the subtotal of the purchase
        temp_subtotal = element[1]['cartSubtotal']
        total_subtotal += temp_subtotal
        temp_row.append(round(temp_subtotal, 2))
        #get the discount applied
        temp_discount =  element[1]['discountTotal']
        total_discount += temp_discount
        temp_row.append(round(temp_discount, 2))
        #get the transport fee
        temp_transport = 0
        try:
            temp_transport = element[1]['orderTransport']
        except KeyError: #older orders did not suppport transport
            pass
        total_transport += temp_transport
        temp_row.append(round(temp_transport, 2))
        #get the tax
        temp_tax = element[1]['cartTaxes']
        total_tax += temp_tax
        temp_row.append(round(temp_tax, 2))
        #get the purchase total
        temp_total = element[0].purchase_total
        total_total += temp_total
        temp_row.append(round(temp_total, 2))
        #obtain the credit period
        temp_row.append(element[0].credit_days)
        #calculate the expiry date of the purchase, parse the invoice date string
        #to a datetime object and add the credit_days
        temp_row.append(getCreditExpiryDate(element[0].invoice_date, element[0].credit_days))
        #get the date on which the invoice was entered in the system
        temp_row.append(str(convertToLocalTimezone(element[0].created)))
        #get the user that entered the purchase in the system
        try:
            
            user_object = json.loads(element[0].user)
            user_first_name = user_object['first_name']
            user_last_name = user_object['last_name']
            if user_first_name == '' and user_last_name == '':
                temp_row.append(user_object['username'])
            else:
                temp_row.append("{} {}".format(user_first_name, user_last_name))
        except Exception as e:
            try:
                cleaned_json = element[0].user.replace("'",'"').replace("True", '"True"').replace('[]', "[]")
                user_object = json.loads(cleaned_json)
                user_first_name = user_object['first_name']
                user_last_name = user_object['last_name']
                if user_first_name == '' and user_last_name == '':
                    temp_row.append(user_object['username'])
                else:
                    temp_row.append("{} {}".format(user_first_name, user_last_name))
            except Exception as e:
                print(e)
                temp_row.append("BAD USER")
        
        data_body.append(temp_row)

    for data_row in data_body:
        for i in range(0, len(data_row)):
            temp_cell = ws.cell(row=current_row, column=current_col+i, value=data_row[i])
        current_row += 1
        current_col = 1
    
    #insert the totals at the totals
    current_col = 4
    current_row += 2

    totals_legend_cell = ws.cell(row=current_row, column=current_col, value="Totales-->")
    totals_legend_cell.font = Font(bold=True, name="Tahoma", size=8)
    current_col+=1

    #add the subtotal
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_subtotal,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1
    #add the total discount
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_discount,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1
    #total transport
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_transport,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1
    #total tax
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_tax,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1
    #total total
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_total,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1



    adjustSheetColsWidth(ws)

def createCashPurchasesReport(purchases, wb, period, sheet_index=1):
    COMPANY_NAME = getCompanyName()
    current_row = 1
    current_col = 1
    header_labels = ['# Consecutivo', '# Factura', 'Fecha Factura', 'Proveedor', 'Subtotal', 'Descuento', 'Transporte', 'Impuesto', 'Total', 'Fecha Ingreso', 'Usuario' ] 
    ws = None
    if sheet_index == 0:
        wb._active_sheet_index = sheet_index
        #get the first sheet, created per default
        ws = wb.active
        #change the title of the sheet
        ws.title = REPORTE_COMPRAS[0]
    else:
        ws = wb.create_sheet(title=REPORTE_COMPRAS[sheet_index])

    #add a simple title
    ws.cell(row=current_row, column=current_col, value=COMPANY_NAME)
    current_row+=1
    #write the date
    now = datetime.now()
    report_generation = "Fecha generación: {}".format(now)
    ws.cell(row=current_row, column=current_col, value=report_generation)
    current_row+=1
    #add the perdiod covered by the report
    report_generation = "Periódo cubierto: {}".format(period)
    ws.cell(row=current_row, column=current_col, value=report_generation)
    current_row+=1
    #write the report type
    ws.cell(row=current_row, column=current_col, value="REPORTE COMPRAS CONTADO")
    current_row+=2

    #write the header
    for i in range(0,len(header_labels)):
        temp_cell = ws.cell(row=current_row, column=current_col+i, value=header_labels[i])
        temp_cell.font = Font(bold=True, name="Tahoma", size=8)
    #reset the col position
    current_col = 1
    current_row+=1
    
    #variable to hold the data of the sheet body
    data_body = []
    
    #track totals for the different columns
    total_subtotal = 0
    total_discount = 0
    total_transport = 0
    total_tax = 0
    total_total = 0
    
    #each element is a tupple with the first index being the purchase
    #and the second on the parsed json cart
    for element in purchases:
        temp_row = []
        #get the internal system consecutive
        temp_row.append(element[0].consecutive)
        #get the invoice number as received
        temp_row.append(element[0].invoice_number)
        #get the date on which the invoice was emitted
        temp_row.append(element[0].invoice_date)
        #get the supplier
        try:
            supplier = json.loads(element[0].supplier)
            temp_row.append("{}-{}".format(supplier['code'], supplier['name']))
        except Exception:
            if element[0].supplier == None or element[0].supplier == '':
                temp_row.append("{}-{}".format('Proveedor', 'Vacío'))
        #get the subtotal of the purchase
        temp_subtotal = element[1]['cartSubtotal']
        total_subtotal += temp_subtotal
        temp_row.append(round(temp_subtotal, 2))
        #get the discount applied
        temp_discount =  element[1]['discountTotal']
        total_discount += temp_discount
        temp_row.append(round(temp_discount, 2))
        #get the transport fee
        temp_transport = 0
        try:
            temp_transport = element[1]['orderTransport']
        except KeyError: #older orders did not suppport transport
            pass
        total_transport += temp_transport
        temp_row.append(round(temp_transport, 2))
        #get the tax
        temp_tax = element[1]['cartTaxes']
        total_tax += temp_tax
        temp_row.append(round(temp_tax, 2))
        #get the purchase total
        temp_total = element[0].purchase_total
        total_total += temp_total
        temp_row.append(round(temp_total, 2))
        #get the date on which the invoice was entered in the system
        temp_row.append(str(convertToLocalTimezone(element[0].created)))
        #get the user that entered the purchase in the system
        try:
            
            user_object = json.loads(element[0].user)
            user_first_name = user_object['first_name']
            user_last_name = user_object['last_name']
            if user_first_name == '' and user_last_name == '':
                temp_row.append(user_object['username'])
            else:
                temp_row.append("{} {}".format(user_first_name, user_last_name))
        except Exception as e:
            try:
                cleaned_json = element[0].user.replace("'",'"').replace("True", '"True"').replace('[]', "[]")
                user_object = json.loads(cleaned_json)
                user_first_name = user_object['first_name']
                user_last_name = user_object['last_name']
                if user_first_name == '' and user_last_name == '':
                    temp_row.append(user_object['username'])
                else:
                    temp_row.append("{} {}".format(user_first_name, user_last_name))
            except Exception as e:
                print(e)
                temp_row.append("BAD USER")
        
        data_body.append(temp_row)

    for data_row in data_body:
        for i in range(0, len(data_row)):
            temp_cell = ws.cell(row=current_row, column=current_col+i, value=data_row[i])
        current_row += 1
        current_col = 1
    
    #insert the totals at the totals
    current_col = 4
    current_row += 2

    totals_legend_cell = ws.cell(row=current_row, column=current_col, value="Totales-->")
    totals_legend_cell.font = Font(bold=True, name="Tahoma", size=8)
    current_col+=1

    #add the subtotal
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_subtotal,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1
    #add the total discount
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_discount,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1
    #total transport
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_transport,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1
    #total tax
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_tax,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1
    #total total
    temp_cell = ws.cell(row=current_row, column=current_col, value=round(total_total,2))
    temp_cell = Font(bold=True, name="Tahoma", size=8)
    current_col +=1



    adjustSheetColsWidth(ws)